import pandas as pd
import numpy as np
import datetime
import os
from openpyxl.utils.cell import coordinate_to_tuple
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# La idea que tengo es la siguiente. 
# En general, la mayoría de las funciones que tengo son para extraer información
# Y el problema que tengo es que siempre le tengo que estar pasando el nombre del archivo
# o algo del estilo. Lo que haré ahora es separar las funciones por tipo.
# Unas serán de extracción de información, y las otras serán de modificación de 
# archivos existes.

class DataExtraction:
    def __init__(self, file, sheet_name=0):
        '''
        Se cargará al objeto los datos que se deseen manipular para extraer información.
        Se tendrá de dos maneras. Una en pd.DataFrame y otra en xl.WorkingBook.

        sheet_name: la hoja que se desea trabajar. Por default es la active sheet.
        '''
        self.df_file = self.__open_excel_file_as_dataframe(file, sheet_name=sheet_name)
        self.wb_file, self.ws_file = self.__open_excel_file_as_working_book(file, sheet_name=sheet_name)

    def __open_excel_file_as_dataframe(self, file, sheet_name=0) -> pd.DataFrame:
        '''
        TODO: Probar que file puede ser una ubicación de archivo y bytes

        Esta función abre el archivo de excel especificado como un pd.DataFrame,
        pero mantiene la misma estructura de las celdas. Las filas empiezan en 1 
        y las columnas empiezan en 1. Así, la esquina superior izquierda (A1) será
        (1,1)
        '''
        df = (pd
         .read_excel(file, skiprows=0, header=None, sheet_name=sheet_name)
         .rename(lambda x: x+1)
         .rename(columns=lambda x: x+1)
        )
        return df
    
    def __open_excel_file_as_working_book(self, file, sheet_name=0) -> (xl.Workbook, xl.Workbook.worksheets):
        '''
        Esta función abre el archivo de excel especificado como un xl.Workbook. Y la hoja deseada.
        '''
        wb = xl.load_workbook(file, data_only=True)
        if sheet_name == 0:
            ws = wb.active
        else:
            # Nos aseguramos de que la hoja que nos dieron esté en el archivo
            assert sheet_name in wb.worksheets, "The given sheet does not exist"
            ws = wb[sheet_name]
        return (wb, ws)
    
    def find_value(self, value_to_find) -> list:
        '''
        Esta función se utiliza para encontrar el value_to_find en todo el excel.
        Usa pandas para que sin importar el tamaño del archivo, el tiempo que se 
        tarde en encontrarlo sea el mismo.

        Regresa una lista de tuplas con las coordenadas convertidas a celdas de excel.
        Ejemplo: si buscas la palabra 'hola' y aparece en la celda A1 y C3 regresa [(1,1), (3,3)]
        Esto es importante recalcarlo porque en pandas empieza en 0, mientras que en excel empieza en 1

        Si el valor no se encuetra, se regresa una lista vacía
        '''

        result = (
            self
            .df_file
            .eq(value_to_find)
            .stack()
            .pipe(lambda df: df[df])
            .index
            .to_list()
        )

        return result

